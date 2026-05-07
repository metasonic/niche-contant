#!/usr/bin/env python3
"""Export dashboard first-page data to an Excel workbook for Google Sheets.

Reuses the same logic from dashboard/app.py to ensure data consistency.
Outputs: dashboard_export.xlsx with multiple sheets, charts, and a dashboard overview.
"""

import csv
import json
import sys
from collections import defaultdict
from pathlib import Path
from statistics import mean, median, stdev

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import (
        BarChart, PieChart, LineChart, Reference, BarChart3D,
    )
    from openpyxl.chart.series import DataPoint
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule
except ImportError:
    print("openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent
EVALUATION_DIR = PROJECT_ROOT / "evaluation"
REPORT_JSON = PROJECT_ROOT / "report.json"
DOWNLOADS_DIR = PROJECT_ROOT / "downloads"
LABELS_PATH = PROJECT_ROOT / "dashboard" / "human_labels.json"

# ---------------------------------------------------------------------------
# Data loading (mirrors dashboard/app.py)
# ---------------------------------------------------------------------------

def discover_csvs():
    if not EVALUATION_DIR.exists():
        return []
    return sorted(EVALUATION_DIR.glob("niche_analysis_results_*.csv"))


def parse_csv(path):
    rows = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            try:
                row["probability"] = float(row.get("probability", 0))
            except (ValueError, TypeError):
                row["probability"] = 0.0
            row["belongs_to_niche"] = row.get("belongs_to_niche", "").strip().upper()
            if row["belongs_to_niche"] not in ("YES", "NO"):
                continue
            rows.append(row)
    return rows


def load_human_labels():
    if LABELS_PATH.exists():
        with open(LABELS_PATH, encoding="utf-8") as f:
            data = json.load(f)
        return data.get("labels", {})
    return {}


def load_report_meta():
    if not REPORT_JSON.exists():
        return {}
    with open(REPORT_JSON, encoding="utf-8") as f:
        report = json.load(f)
    return {
        "generated_at": report.get("generated_at", ""),
        "total_niches": report.get("total_niches", 0),
        "grand_total_images": report.get("grand_total_images", 0),
        "grand_total_posts": report.get("grand_total_posts", 0),
        "grand_total_videos_downloaded": report.get("grand_total_videos_downloaded", 0),
    }


# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=11, color="2F5496")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
KPI_VALUE_FONT = Font(bold=True, size=13, color="2F5496")
KPI_LABEL_FONT = Font(size=10, color="595959")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# Chart color palette
COLORS = [
    "4472C4", "ED7D31", "A5A5A5", "FFC000", "5B9BD5",
    "70AD47", "264478", "9B57A0", "636363", "EB7E30",
]


def style_header_row(ws, row=1, max_col=None):
    max_col = max_col or ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row=2, max_col=None):
    max_col = max_col or ws.max_column
    alt_fill = PatternFill(start_color="F2F7FB", end_color="F2F7FB", fill_type="solid")
    for row_idx in range(start_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")
            if row_idx % 2 == 0:
                cell.fill = alt_fill


def auto_width(ws, min_width=10, max_width=40):
    for col_cells in ws.columns:
        lengths = []
        for cell in col_cells:
            if cell.value is not None:
                lengths.append(len(str(cell.value)))
        if lengths:
            width = min(max(max(lengths) + 2, min_width), max_width)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = width


def apply_heatmap_colors(ws, min_row, max_row, min_col, max_col):
    """Apply 3-color scale conditional formatting (red-yellow-green)."""
    cell_range = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    ws.conditional_formatting.add(
        cell_range,
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        ),
    )


# ---------------------------------------------------------------------------
# Main export
# ---------------------------------------------------------------------------

def main():
    csv_paths = discover_csvs()
    if not csv_paths:
        print("No evaluation CSVs found in evaluation/")
        sys.exit(1)

    all_rows = []
    for p in csv_paths:
        all_rows.extend(parse_csv(p))

    if not all_rows:
        print("No valid evaluation rows found.")
        sys.exit(1)

    models = sorted({r["model_name_name"] for r in all_rows})
    niches = sorted({r["folder_name"] for r in all_rows})
    model_short = {m: m.split("/")[-1] for m in models}
    short_names = [model_short[m] for m in models]
    all_labels = load_human_labels()
    report_meta = load_report_meta()

    total_rows_count = len(all_rows)
    total_yes = sum(1 for r in all_rows if r["belongs_to_niche"] == "YES")
    all_probs = [r["probability"] for r in all_rows]

    # Image-level agreement
    image_verdicts = defaultdict(dict)
    for r in all_rows:
        key = f"{r['folder_name']}/{r['image_name']}"
        image_verdicts[key][model_short[r["model_name_name"]]] = r["belongs_to_niche"]

    total_images_eval = len(image_verdicts)
    unanimous_yes = unanimous_no = disagreements_count = 0
    disagreement_samples = []
    for img_key, verdicts in image_verdicts.items():
        vals = set(verdicts.values())
        if len(vals) == 1:
            if "YES" in vals:
                unanimous_yes += 1
            else:
                unanimous_no += 1
        else:
            disagreements_count += 1
            if len(disagreement_samples) < 20:
                disagreement_samples.append({"image": img_key, "verdicts": verdicts})
    agreement_rate = round((unanimous_yes + unanimous_no) / total_images_eval * 100, 1) if total_images_eval else 0

    # Pre-compute per-model data
    per_model_data = {}
    for m in models:
        short = model_short[m]
        m_rows = [r for r in all_rows if r["model_name_name"] == m]
        total = len(m_rows)
        yes_count = sum(1 for r in m_rows if r["belongs_to_niche"] == "YES")
        probs = [r["probability"] for r in m_rows]
        per_model_data[short] = {
            "full_name": m, "total": total, "yes": yes_count, "no": total - yes_count,
            "yes_rate": round(yes_count / total * 100, 1) if total else 0,
            "avg_conf": round(mean(probs), 3) if probs else 0,
            "med_conf": round(median(probs), 3) if probs else 0,
            "std_conf": round(stdev(probs), 3) if len(probs) > 1 else 0,
        }

    # Pre-compute per-niche data
    per_niche_data = {}
    for n in niches:
        n_rows = [r for r in all_rows if r["folder_name"] == n]
        total = len(n_rows)
        yes_count = sum(1 for r in n_rows if r["belongs_to_niche"] == "YES")
        probs = [r["probability"] for r in n_rows]
        per_niche_data[n] = {
            "total": total, "yes": yes_count, "no": total - yes_count,
            "yes_rate": round(yes_count / total * 100, 1) if total else 0,
            "avg_conf": round(mean(probs), 3) if probs else 0,
        }

    wb = openpyxl.Workbook()

    # ===================================================================
    # SHEET 1: DASHBOARD OVERVIEW (main sheet with KPIs + charts)
    # ===================================================================
    ws_dash = wb.active
    ws_dash.title = "Dashboard Overview"
    ws_dash.sheet_properties.tabColor = "2F5496"

    # Title
    ws_dash.merge_cells("A1:H1")
    title_cell = ws_dash["A1"]
    title_cell.value = "Vision Model Evaluation — Dashboard Overview"
    title_cell.font = Font(bold=True, size=18, color="1F3864")
    title_cell.alignment = Alignment(horizontal="center")
    ws_dash.row_dimensions[1].height = 35

    # Subtitle
    ws_dash.merge_cells("A2:H2")
    sub_cell = ws_dash["A2"]
    gen_at = report_meta.get("generated_at", "N/A")
    sub_cell.value = f"Generated: {gen_at}  |  {len(models)} Models  |  {len(niches)} Niches  |  {total_images_eval} Images  |  {total_rows_count} Evaluations"
    sub_cell.font = Font(size=10, color="808080")
    sub_cell.alignment = Alignment(horizontal="center")

    # --- KPI Cards (row 4-6) ---
    kpi_items = [
        ("Total Models", len(models)),
        ("Total Niches", len(niches)),
        ("Images Evaluated", total_images_eval),
        ("Total Evaluations", total_rows_count),
        ("YES Rate", f"{round(total_yes / total_rows_count * 100, 1)}%"),
        ("Avg Confidence", f"{round(mean(all_probs), 3)}"),
        ("Agreement Rate", f"{agreement_rate}%"),
        ("Disagreements", disagreements_count),
    ]
    kpi_fill = PatternFill(start_color="E8EEF7", end_color="E8EEF7", fill_type="solid")
    kpi_border = Border(
        left=Side(style="medium", color="2F5496"),
        right=Side(style="medium", color="2F5496"),
        top=Side(style="medium", color="2F5496"),
        bottom=Side(style="medium", color="2F5496"),
    )
    for idx, (label, value) in enumerate(kpi_items):
        col = idx + 1
        # Value row
        cell_v = ws_dash.cell(row=4, column=col, value=value)
        cell_v.font = KPI_VALUE_FONT
        cell_v.alignment = Alignment(horizontal="center")
        cell_v.fill = kpi_fill
        cell_v.border = kpi_border
        # Label row
        cell_l = ws_dash.cell(row=5, column=col, value=label)
        cell_l.font = KPI_LABEL_FONT
        cell_l.alignment = Alignment(horizontal="center")
        cell_l.fill = kpi_fill
        cell_l.border = kpi_border
        ws_dash.column_dimensions[get_column_letter(col)].width = 18

    # --- Embedded data tables for charts (hidden off to the right) ---
    # Model YES Rate data starting at K8
    data_start_col = 11  # column K
    ws_dash.cell(row=7, column=data_start_col, value="Model").font = HEADER_FONT_WHITE
    ws_dash.cell(row=7, column=data_start_col + 1, value="YES Rate (%)").font = HEADER_FONT_WHITE
    ws_dash.cell(row=7, column=data_start_col + 2, value="Avg Confidence").font = HEADER_FONT_WHITE
    ws_dash.cell(row=7, column=data_start_col + 3, value="YES Count").font = HEADER_FONT_WHITE
    ws_dash.cell(row=7, column=data_start_col + 4, value="NO Count").font = HEADER_FONT_WHITE
    for ci in range(5):
        c = ws_dash.cell(row=7, column=data_start_col + ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT_WHITE

    for i, short in enumerate(short_names):
        row = 8 + i
        d = per_model_data[short]
        ws_dash.cell(row=row, column=data_start_col, value=short)
        ws_dash.cell(row=row, column=data_start_col + 1, value=d["yes_rate"])
        ws_dash.cell(row=row, column=data_start_col + 2, value=d["avg_conf"])
        ws_dash.cell(row=row, column=data_start_col + 3, value=d["yes"])
        ws_dash.cell(row=row, column=data_start_col + 4, value=d["no"])

    num_models = len(short_names)

    # Niche data starting below model data
    niche_data_row = 8 + num_models + 2
    ws_dash.cell(row=niche_data_row, column=data_start_col, value="Niche").font = HEADER_FONT_WHITE
    ws_dash.cell(row=niche_data_row, column=data_start_col + 1, value="YES Rate (%)").font = HEADER_FONT_WHITE
    ws_dash.cell(row=niche_data_row, column=data_start_col + 2, value="Avg Confidence").font = HEADER_FONT_WHITE
    ws_dash.cell(row=niche_data_row, column=data_start_col + 3, value="YES Count").font = HEADER_FONT_WHITE
    ws_dash.cell(row=niche_data_row, column=data_start_col + 4, value="NO Count").font = HEADER_FONT_WHITE
    for ci in range(5):
        c = ws_dash.cell(row=niche_data_row, column=data_start_col + ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT_WHITE

    for i, n in enumerate(niches):
        row = niche_data_row + 1 + i
        d = per_niche_data[n]
        ws_dash.cell(row=row, column=data_start_col, value=n)
        ws_dash.cell(row=row, column=data_start_col + 1, value=d["yes_rate"])
        ws_dash.cell(row=row, column=data_start_col + 2, value=d["avg_conf"])
        ws_dash.cell(row=row, column=data_start_col + 3, value=d["yes"])
        ws_dash.cell(row=row, column=data_start_col + 4, value=d["no"])

    num_niches = len(niches)

    # Agreement pie data
    agree_data_row = niche_data_row + num_niches + 3
    ws_dash.cell(row=agree_data_row, column=data_start_col, value="Category").font = HEADER_FONT_WHITE
    ws_dash.cell(row=agree_data_row, column=data_start_col + 1, value="Count").font = HEADER_FONT_WHITE
    for ci in range(2):
        c = ws_dash.cell(row=agree_data_row, column=data_start_col + ci)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT_WHITE
    ws_dash.cell(row=agree_data_row + 1, column=data_start_col, value="Unanimous YES")
    ws_dash.cell(row=agree_data_row + 1, column=data_start_col + 1, value=unanimous_yes)
    ws_dash.cell(row=agree_data_row + 2, column=data_start_col, value="Unanimous NO")
    ws_dash.cell(row=agree_data_row + 2, column=data_start_col + 1, value=unanimous_no)
    ws_dash.cell(row=agree_data_row + 3, column=data_start_col, value="Disagreement")
    ws_dash.cell(row=agree_data_row + 3, column=data_start_col + 1, value=disagreements_count)

    # --- CHART 1: Model YES Rate (bar chart) → placed at A7 ---
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "YES Rate by Model (%)"
    chart1.y_axis.title = "YES Rate (%)"
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100
    chart1.x_axis.title = "Model"
    data_ref = Reference(ws_dash, min_col=data_start_col + 1, min_row=7, max_row=7 + num_models)
    cats_ref = Reference(ws_dash, min_col=data_start_col, min_row=8, max_row=7 + num_models)
    chart1.add_data(data_ref, titles_from_data=True)
    chart1.set_categories(cats_ref)
    chart1.shape = 4
    chart1.width = 22
    chart1.height = 13
    chart1.legend = None
    # Color each bar
    for i in range(num_models):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = COLORS[i % len(COLORS)]
        chart1.series[0].data_points.append(pt)
    ws_dash.add_chart(chart1, "A7")

    # --- CHART 2: Agreement Pie → placed at A24 ---
    pie = PieChart()
    pie.title = "Model Agreement Distribution"
    pie.style = 10
    pie_data = Reference(ws_dash, min_col=data_start_col + 1, min_row=agree_data_row, max_row=agree_data_row + 3)
    pie_cats = Reference(ws_dash, min_col=data_start_col, min_row=agree_data_row + 1, max_row=agree_data_row + 3)
    pie.add_data(pie_data, titles_from_data=True)
    pie.set_categories(pie_cats)
    pie.width = 16
    pie.height = 13
    # Colors for pie slices
    pie_colors = ["63BE7B", "F8696B", "FFEB84"]
    for i, color in enumerate(pie_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie.series[0].data_points.append(pt)
    pie.series[0].dLbls = DataLabelList()
    pie.series[0].dLbls.showPercent = True
    pie.series[0].dLbls.showVal = True
    ws_dash.add_chart(pie, "E7")

    # --- CHART 3: Niche YES Rate (horizontal bar) → placed at A24 ---
    chart3 = BarChart()
    chart3.type = "bar"
    chart3.style = 10
    chart3.title = "YES Rate by Niche (%)"
    chart3.x_axis.title = "YES Rate (%)"
    chart3.x_axis.scaling.min = 0
    chart3.x_axis.scaling.max = 100
    niche_data_ref = Reference(ws_dash, min_col=data_start_col + 1, min_row=niche_data_row, max_row=niche_data_row + num_niches)
    niche_cats_ref = Reference(ws_dash, min_col=data_start_col, min_row=niche_data_row + 1, max_row=niche_data_row + num_niches)
    chart3.add_data(niche_data_ref, titles_from_data=True)
    chart3.set_categories(niche_cats_ref)
    chart3.shape = 4
    chart3.width = 22
    chart3.height = 14
    chart3.legend = None
    for i in range(num_niches):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = COLORS[i % len(COLORS)]
        chart3.series[0].data_points.append(pt)
    ws_dash.add_chart(chart3, "A24")

    # --- CHART 4: Model Avg Confidence (bar chart) → placed at E24 ---
    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = "Average Confidence by Model"
    chart4.y_axis.title = "Avg Confidence"
    chart4.y_axis.scaling.min = 0
    chart4.y_axis.scaling.max = 1.0
    conf_ref = Reference(ws_dash, min_col=data_start_col + 2, min_row=7, max_row=7 + num_models)
    chart4.add_data(conf_ref, titles_from_data=True)
    chart4.set_categories(cats_ref)
    chart4.width = 22
    chart4.height = 14
    chart4.legend = None
    for i in range(num_models):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = COLORS[i % len(COLORS)]
        chart4.series[0].data_points.append(pt)
    ws_dash.add_chart(chart4, "E24")

    # --- Summary table below charts (row ~42) ---
    summary_row = 42
    ws_dash.merge_cells(f"A{summary_row}:H{summary_row}")
    ws_dash.cell(row=summary_row, column=1, value="Per-Model Summary").font = TITLE_FONT

    headers_m = ["Model", "Total", "YES", "NO", "YES Rate (%)", "Avg Conf", "Med Conf", "Std Conf"]
    for ci, h in enumerate(headers_m, 1):
        ws_dash.cell(row=summary_row + 1, column=ci, value=h)
    style_header_row(ws_dash, row=summary_row + 1, max_col=8)

    for i, short in enumerate(short_names):
        row = summary_row + 2 + i
        d = per_model_data[short]
        ws_dash.cell(row=row, column=1, value=short)
        ws_dash.cell(row=row, column=2, value=d["total"])
        ws_dash.cell(row=row, column=3, value=d["yes"])
        ws_dash.cell(row=row, column=4, value=d["no"])
        ws_dash.cell(row=row, column=5, value=d["yes_rate"])
        ws_dash.cell(row=row, column=6, value=d["avg_conf"])
        ws_dash.cell(row=row, column=7, value=d["med_conf"])
        ws_dash.cell(row=row, column=8, value=d["std_conf"])
    style_data_rows(ws_dash, start_row=summary_row + 2, max_col=8)

    niche_summary_row = summary_row + 2 + num_models + 2
    ws_dash.merge_cells(f"A{niche_summary_row}:F{niche_summary_row}")
    ws_dash.cell(row=niche_summary_row, column=1, value="Per-Niche Summary").font = TITLE_FONT

    headers_n = ["Niche", "Total", "YES", "NO", "YES Rate (%)", "Avg Conf"]
    for ci, h in enumerate(headers_n, 1):
        ws_dash.cell(row=niche_summary_row + 1, column=ci, value=h)
    style_header_row(ws_dash, row=niche_summary_row + 1, max_col=6)

    for i, n in enumerate(niches):
        row = niche_summary_row + 2 + i
        d = per_niche_data[n]
        ws_dash.cell(row=row, column=1, value=n)
        ws_dash.cell(row=row, column=2, value=d["total"])
        ws_dash.cell(row=row, column=3, value=d["yes"])
        ws_dash.cell(row=row, column=4, value=d["no"])
        ws_dash.cell(row=row, column=5, value=d["yes_rate"])
        ws_dash.cell(row=row, column=6, value=d["avg_conf"])
    style_data_rows(ws_dash, start_row=niche_summary_row + 2, max_col=6)

    # ===================================================================
    # SHEET 2: Per-Model Metrics + Chart
    # ===================================================================
    ws2 = wb.create_sheet("Per-Model Metrics")
    ws2.sheet_properties.tabColor = "ED7D31"
    headers = ["Model", "Full Name", "Total Evals", "YES Count", "NO Count",
               "YES Rate (%)", "Avg Confidence", "Median Confidence", "Std Confidence"]
    ws2.append(headers)
    for short in short_names:
        d = per_model_data[short]
        ws2.append([short, d["full_name"], d["total"], d["yes"], d["no"],
                     d["yes_rate"], d["avg_conf"], d["med_conf"], d["std_conf"]])
    style_header_row(ws2)
    style_data_rows(ws2)
    auto_width(ws2)

    # Stacked bar: YES vs NO per model
    chart_m = BarChart()
    chart_m.type = "col"
    chart_m.grouping = "stacked"
    chart_m.title = "YES vs NO Count by Model"
    chart_m.y_axis.title = "Count"
    yes_ref = Reference(ws2, min_col=4, min_row=1, max_row=1 + num_models)
    no_ref = Reference(ws2, min_col=5, min_row=1, max_row=1 + num_models)
    cats = Reference(ws2, min_col=1, min_row=2, max_row=1 + num_models)
    chart_m.add_data(yes_ref, titles_from_data=True)
    chart_m.add_data(no_ref, titles_from_data=True)
    chart_m.set_categories(cats)
    chart_m.series[0].graphicalProperties.solidFill = "63BE7B"
    chart_m.series[1].graphicalProperties.solidFill = "F8696B"
    chart_m.width = 24
    chart_m.height = 14
    ws2.add_chart(chart_m, "A" + str(num_models + 4))

    # Confidence comparison chart
    chart_conf = BarChart()
    chart_conf.type = "col"
    chart_conf.title = "Confidence Statistics by Model"
    chart_conf.y_axis.title = "Confidence"
    for col_idx, label in [(7, "Avg"), (8, "Median")]:
        ref = Reference(ws2, min_col=col_idx, min_row=1, max_row=1 + num_models)
        chart_conf.add_data(ref, titles_from_data=True)
    chart_conf.set_categories(cats)
    chart_conf.width = 24
    chart_conf.height = 14
    chart_conf.series[0].graphicalProperties.solidFill = "4472C4"
    chart_conf.series[1].graphicalProperties.solidFill = "FFC000"
    ws2.add_chart(chart_conf, "A" + str(num_models + 22))

    # ===================================================================
    # SHEET 3: Per-Niche Metrics + Chart
    # ===================================================================
    ws3 = wb.create_sheet("Per-Niche Metrics")
    ws3.sheet_properties.tabColor = "70AD47"
    headers = ["Niche", "Total Evals", "YES Count", "NO Count", "YES Rate (%)", "Avg Confidence"]
    ws3.append(headers)
    for n in niches:
        d = per_niche_data[n]
        ws3.append([n, d["total"], d["yes"], d["no"], d["yes_rate"], d["avg_conf"]])
    style_header_row(ws3)
    style_data_rows(ws3)
    auto_width(ws3)

    # Stacked bar: YES vs NO per niche
    chart_n = BarChart()
    chart_n.type = "bar"
    chart_n.grouping = "stacked"
    chart_n.title = "YES vs NO Count by Niche"
    chart_n.x_axis.title = "Count"
    yes_n = Reference(ws3, min_col=3, min_row=1, max_row=1 + num_niches)
    no_n = Reference(ws3, min_col=4, min_row=1, max_row=1 + num_niches)
    cats_n = Reference(ws3, min_col=1, min_row=2, max_row=1 + num_niches)
    chart_n.add_data(yes_n, titles_from_data=True)
    chart_n.add_data(no_n, titles_from_data=True)
    chart_n.set_categories(cats_n)
    chart_n.series[0].graphicalProperties.solidFill = "63BE7B"
    chart_n.series[1].graphicalProperties.solidFill = "F8696B"
    chart_n.width = 24
    chart_n.height = 16
    ws3.add_chart(chart_n, "A" + str(num_niches + 4))

    # ===================================================================
    # SHEET 4: YES Rate Heatmap (Model x Niche) + conditional formatting
    # ===================================================================
    ws4 = wb.create_sheet("Heatmap - YES Rate")
    ws4.sheet_properties.tabColor = "FFC000"
    headers = ["Model"] + niches
    ws4.append(headers)
    for m in models:
        short = model_short[m]
        row_data = [short]
        for n in niches:
            mn_rows = [r for r in all_rows if r["model_name_name"] == m and r["folder_name"] == n]
            total = len(mn_rows)
            yes_c = sum(1 for r in mn_rows if r["belongs_to_niche"] == "YES")
            row_data.append(round(yes_c / total * 100, 1) if total else 0)
        ws4.append(row_data)
    style_header_row(ws4)
    style_data_rows(ws4)
    auto_width(ws4)
    apply_heatmap_colors(ws4, min_row=2, max_row=1 + num_models, min_col=2, max_col=1 + num_niches)

    # Grouped bar chart for heatmap data
    chart_hm = BarChart()
    chart_hm.type = "col"
    chart_hm.grouping = "clustered"
    chart_hm.title = "YES Rate (%) — Model x Niche"
    chart_hm.y_axis.title = "YES Rate (%)"
    chart_hm.y_axis.scaling.max = 100
    for i in range(num_models):
        ref = Reference(ws4, min_col=2, max_col=1 + num_niches, min_row=2 + i, max_row=2 + i)
        chart_hm.add_data(ref, from_rows=True, titles_from_data=False)
        chart_hm.series[i].title = openpyxl.chart.series.SeriesLabel(v=short_names[i])
        chart_hm.series[i].graphicalProperties.solidFill = COLORS[i % len(COLORS)]
    cats_hm = Reference(ws4, min_col=2, max_col=1 + num_niches, min_row=1)
    chart_hm.set_categories(cats_hm)
    chart_hm.width = 32
    chart_hm.height = 16
    ws4.add_chart(chart_hm, "A" + str(num_models + 4))

    # ===================================================================
    # SHEET 5: Confidence Heatmap + conditional formatting
    # ===================================================================
    ws5 = wb.create_sheet("Heatmap - Confidence")
    ws5.sheet_properties.tabColor = "5B9BD5"
    headers = ["Model"] + niches
    ws5.append(headers)
    for m in models:
        short = model_short[m]
        row_data = [short]
        for n in niches:
            mn_rows = [r for r in all_rows if r["model_name_name"] == m and r["folder_name"] == n]
            probs = [r["probability"] for r in mn_rows]
            row_data.append(round(mean(probs), 3) if probs else 0)
        ws5.append(row_data)
    style_header_row(ws5)
    style_data_rows(ws5)
    auto_width(ws5)
    apply_heatmap_colors(ws5, min_row=2, max_row=1 + num_models, min_col=2, max_col=1 + num_niches)

    # ===================================================================
    # SHEET 6: Confidence Distribution + Chart
    # ===================================================================
    ws6 = wb.create_sheet("Confidence Distribution")
    ws6.sheet_properties.tabColor = "9B57A0"
    bucket_labels = ["0.0-0.2", "0.2-0.4", "0.4-0.6", "0.6-0.8", "0.8-1.0"]
    bucket_ranges = [(0, 0.2), (0.2, 0.4), (0.4, 0.6), (0.6, 0.8), (0.8, 1.01)]
    headers = ["Model"] + bucket_labels
    ws6.append(headers)
    for m in models:
        short = model_short[m]
        m_rows = [r for r in all_rows if r["model_name_name"] == m]
        probs = [r["probability"] for r in m_rows]
        counts = [sum(1 for p in probs if lo <= p < hi) for lo, hi in bucket_ranges]
        ws6.append([short] + counts)
    style_header_row(ws6)
    style_data_rows(ws6)
    auto_width(ws6)

    # Stacked bar: confidence distribution per model
    chart_cd = BarChart()
    chart_cd.type = "col"
    chart_cd.grouping = "stacked"
    chart_cd.title = "Confidence Distribution by Model"
    chart_cd.y_axis.title = "Count"
    bucket_colors = ["F8696B", "FFAA5C", "FFEB84", "A8D08D", "63BE7B"]
    for bi in range(5):
        ref = Reference(ws6, min_col=2 + bi, min_row=1, max_row=1 + num_models)
        chart_cd.add_data(ref, titles_from_data=True)
        chart_cd.series[bi].graphicalProperties.solidFill = bucket_colors[bi]
    cats_cd = Reference(ws6, min_col=1, min_row=2, max_row=1 + num_models)
    chart_cd.set_categories(cats_cd)
    chart_cd.width = 28
    chart_cd.height = 16
    ws6.add_chart(chart_cd, "A" + str(num_models + 4))

    # ===================================================================
    # SHEET 7: Model Agreement Per Image
    # ===================================================================
    ws7 = wb.create_sheet("Agreement Per Image")
    ws7.sheet_properties.tabColor = "636363"
    model_names_short = sorted(model_short.values())
    headers = ["Image Key", "Niche"] + model_names_short + ["Agreement Status", "YES Votes", "NO Votes"]
    ws7.append(headers)
    for img_key, verdicts in sorted(image_verdicts.items()):
        niche = img_key.split("/", 1)[0]
        row_data = [img_key, niche]
        yes_votes = no_votes = 0
        for ms in model_names_short:
            v = verdicts.get(ms, "")
            row_data.append(v)
            if v == "YES":
                yes_votes += 1
            elif v == "NO":
                no_votes += 1
        vals = set(verdicts.values())
        if len(vals) == 1:
            status = "Unanimous YES" if "YES" in vals else "Unanimous NO"
        else:
            status = "Disagreement"
        row_data.extend([status, yes_votes, no_votes])
        ws7.append(row_data)
    style_header_row(ws7)
    auto_width(ws7)

    # Agreement summary mini-table + pie chart on the same sheet
    ag_summary_row = total_images_eval + 4
    ws7.cell(row=ag_summary_row, column=1, value="Agreement Summary").font = TITLE_FONT
    ws7.cell(row=ag_summary_row + 1, column=1, value="Category")
    ws7.cell(row=ag_summary_row + 1, column=2, value="Count")
    style_header_row(ws7, row=ag_summary_row + 1, max_col=2)
    ws7.cell(row=ag_summary_row + 2, column=1, value="Unanimous YES")
    ws7.cell(row=ag_summary_row + 2, column=2, value=unanimous_yes)
    ws7.cell(row=ag_summary_row + 3, column=1, value="Unanimous NO")
    ws7.cell(row=ag_summary_row + 3, column=2, value=unanimous_no)
    ws7.cell(row=ag_summary_row + 4, column=1, value="Disagreement")
    ws7.cell(row=ag_summary_row + 4, column=2, value=disagreements_count)

    pie2 = PieChart()
    pie2.title = "Agreement Breakdown"
    pie2_data = Reference(ws7, min_col=2, min_row=ag_summary_row + 1, max_row=ag_summary_row + 4)
    pie2_cats = Reference(ws7, min_col=1, min_row=ag_summary_row + 2, max_row=ag_summary_row + 4)
    pie2.add_data(pie2_data, titles_from_data=True)
    pie2.set_categories(pie2_cats)
    pie2.width = 16
    pie2.height = 13
    for i, color in enumerate(pie_colors):
        pt = DataPoint(idx=i)
        pt.graphicalProperties.solidFill = color
        pie2.series[0].data_points.append(pt)
    pie2.series[0].dLbls = DataLabelList()
    pie2.series[0].dLbls.showPercent = True
    ws7.add_chart(pie2, "D" + str(ag_summary_row))

    # ===================================================================
    # SHEET 8: Low Confidence Items
    # ===================================================================
    ws8 = wb.create_sheet("Low Confidence Items")
    ws8.sheet_properties.tabColor = "F8696B"
    headers = ["Model", "Niche", "Image", "Verdict", "Probability", "Reason"]
    ws8.append(headers)
    sorted_by_conf = sorted(all_rows, key=lambda r: r["probability"])
    for r in sorted_by_conf[:30]:
        ws8.append([
            model_short[r["model_name_name"]], r["folder_name"], r["image_name"],
            r["belongs_to_niche"], r["probability"], r.get("reason", ""),
        ])
    style_header_row(ws8)
    style_data_rows(ws8)
    auto_width(ws8)

    # ===================================================================
    # SHEET 9: Raw Data
    # ===================================================================
    ws9 = wb.create_sheet("Raw Data")
    ws9.sheet_properties.tabColor = "A5A5A5"
    headers = ["Model (Short)", "Model (Full)", "Niche", "Image", "Verdict", "Probability", "Reason"]
    ws9.append(headers)
    for r in all_rows:
        ws9.append([
            model_short[r["model_name_name"]], r["model_name_name"],
            r["folder_name"], r["image_name"],
            r["belongs_to_niche"], r["probability"], r.get("reason", ""),
        ])
    style_header_row(ws9)
    auto_width(ws9)

    # ===================================================================
    # SHEET 10: Ground Truth (if human labels exist)
    # ===================================================================
    if all_labels:
        ws10 = wb.create_sheet("Ground Truth - Model Accuracy")
        ws10.sheet_properties.tabColor = "264478"
        headers = ["Model", "TP", "FP", "TN", "FN", "Total Labeled",
                   "Accuracy (%)", "Precision (%)", "Recall (%)", "F1 (%)",
                   "Specificity (%)", "Cohen's Kappa", "ECE"]
        ws10.append(headers)

        gt_rows = []
        for m in models:
            short = model_short[m]
            tp = fp = tn = fn = 0
            m_rows = [r for r in all_rows if r["model_name_name"] == m]
            for r in m_rows:
                key = f"{r['folder_name']}/{r['image_name']}"
                lbl_entry = all_labels.get(key)
                if not lbl_entry:
                    continue
                gt = lbl_entry["verdict"]
                pred = r["belongs_to_niche"]
                if pred == "YES" and gt == "YES": tp += 1
                elif pred == "YES" and gt == "NO": fp += 1
                elif pred == "NO" and gt == "NO": tn += 1
                elif pred == "NO" and gt == "YES": fn += 1
            total_labeled = tp + fp + tn + fn
            accuracy = round((tp + tn) / total_labeled * 100, 1) if total_labeled else None
            precision = round(tp / (tp + fp) * 100, 1) if (tp + fp) else None
            recall = round(tp / (tp + fn) * 100, 1) if (tp + fn) else None
            f1 = round(2 * tp / (2 * tp + fp + fn) * 100, 1) if (2 * tp + fp + fn) else None
            specificity = round(tn / (tn + fp) * 100, 1) if (tn + fp) > 0 else None
            if total_labeled > 0:
                p_o = (tp + tn) / total_labeled
                p_yes = ((tp + fp) / total_labeled) * ((tp + fn) / total_labeled)
                p_no = ((tn + fn) / total_labeled) * ((tn + fp) / total_labeled)
                p_e = p_yes + p_no
                kappa = round((p_o - p_e) / (1 - p_e), 3) if p_e != 1 else None
            else:
                kappa = None
            cal_bounds = [(0, 0.2), (0.2, 0.4), (0.4, 0.6), (0.6, 0.8), (0.8, 1.01)]
            ece_num = ece_den = 0
            for lo, hi in cal_bounds:
                bucket_rows = []
                for r in m_rows:
                    key = f"{r['folder_name']}/{r['image_name']}"
                    le = all_labels.get(key)
                    if not le: continue
                    prob = r["probability"]
                    if lo <= prob < hi:
                        bucket_rows.append((prob, 1 if r["belongs_to_niche"] == le["verdict"] else 0))
                if bucket_rows:
                    avg_conf = mean([x[0] for x in bucket_rows])
                    actual_acc = sum(x[1] for x in bucket_rows) / len(bucket_rows)
                    ece_num += len(bucket_rows) * abs(avg_conf - actual_acc)
                    ece_den += len(bucket_rows)
            ece = round(ece_num / ece_den * 100, 1) if ece_den > 0 else None
            gt_rows.append([short, tp, fp, tn, fn, total_labeled,
                           accuracy, precision, recall, f1, specificity, kappa, ece])
            ws10.append(gt_rows[-1])
        style_header_row(ws10)
        style_data_rows(ws10)
        auto_width(ws10)

        # Chart: Accuracy/Precision/Recall/F1 per model
        if gt_rows:
            chart_gt = BarChart()
            chart_gt.type = "col"
            chart_gt.grouping = "clustered"
            chart_gt.title = "Ground Truth Metrics by Model"
            chart_gt.y_axis.title = "Percentage (%)"
            for ci, label in [(7, "Accuracy"), (8, "Precision"), (9, "Recall"), (10, "F1")]:
                ref = Reference(ws10, min_col=ci, min_row=1, max_row=1 + len(gt_rows))
                chart_gt.add_data(ref, titles_from_data=True)
            cats_gt = Reference(ws10, min_col=1, min_row=2, max_row=1 + len(gt_rows))
            chart_gt.set_categories(cats_gt)
            chart_gt.width = 28
            chart_gt.height = 16
            gt_colors = ["4472C4", "70AD47", "FFC000", "ED7D31"]
            for i, c in enumerate(gt_colors):
                chart_gt.series[i].graphicalProperties.solidFill = c
            ws10.add_chart(chart_gt, "A" + str(len(gt_rows) + 4))

        ws11 = wb.create_sheet("Human Labels")
        ws11.append(["Image Key", "Verdict", "Timestamp"])
        for key, entry in sorted(all_labels.items()):
            ws11.append([key, entry["verdict"], entry.get("timestamp", "")])
        style_header_row(ws11)
        auto_width(ws11)

    # ===================================================================
    # Save
    # ===================================================================
    output_path = PROJECT_ROOT / "dashboard_export.xlsx"
    wb.save(str(output_path))
    print(f"Exported to: {output_path}")
    print(f"Sheets: {wb.sheetnames}")
    print(f"Total raw evaluation rows: {len(all_rows)}")
    print(f"Models: {len(models)}, Niches: {len(niches)}, Images: {total_images_eval}")
    print(f"Charts embedded: ~10 (bars, pies, heatmaps)")


if __name__ == "__main__":
    main()
